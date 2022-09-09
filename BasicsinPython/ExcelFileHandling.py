import openpyxl


class ExcelfileHandling():

    def readdata(self):
        # Open the excel file
        excelopen = openpyxl.load_workbook(
            "C:\\Users\\sathishkumar\\PycharmProjects\\Python7AMBesantOnline\\Inputfile\\Login.xlsx")
        #excelopenoutput = openpyxl.load_workbook("C:\\Users\\sathishkumar\\PycharmProjects\\Python7AMBesantOnline\\Inputfile\\output.xlsx","w")
        sheetname = excelopen["Credentials"]
        excelopenoutput = openpyxl.Workbook() # output workbook
        sheetoutputname = excelopenoutput.active # output sheet
        cell = sheetname.cell(row=1, column=3).value
        print(cell)
        maximumrow = sheetname.max_row
        print(maximumrow)
        maximumcolumn = sheetname.max_column
        print(maximumcolumn)

        for rowvalue in range(1, maximumrow + 1): #row
            for columnvalue in range(1, maximumcolumn + 1): #column
                print(sheetname.cell(row=rowvalue, column=columnvalue).value)
                eachvalue=sheetname.cell(row=rowvalue, column=columnvalue).value
                sheetoutputname.cell(row=rowvalue, column=columnvalue).value = eachvalue
                #sheetoutputname.cell(row=rowvalue, column=columnvalue).value = "entred through script"
        excelopenoutput.save("C:\\Users\\sathishkumar\\PycharmProjects\\Python7AMBesantOnline\\Inputfile\\output2.xlsx")
        excelopenoutput.close()


obj = ExcelfileHandling()
obj.readdata()
