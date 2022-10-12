
import openpyxl as openpyxl

class Excel_Read:
    Dict ={}

    @staticmethod
    def retrundicvalue():
        Dict = {}
        book = openpyxl.load_workbook(
            "C:\\Users\\sathishkumar\\PycharmProjects\\Pythongitupload\\Inputfile\\Facebooklogin.xlsx")
        sheet = book["Credentials"]
        cell = sheet.cell(row=1, column=2)

        print(cell)
        max_row = sheet.max_row
        print(max_row)

        # for max column
        max_column = sheet.max_column
        print(max_column)

        for i in range(2, max_row + 1):
            for j in range(1, max_column + 1):
                Dict[(sheet.cell(row=1, column=j).value) + str(i - 1)] = sheet.cell(row=i, column=j).value
                #Dict[(sheet.cell(row=1, column=j).value)] = sheet.cell(row=i, column=j).value

        return Dict